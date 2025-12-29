#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
factbank_ingest.py

PDF 리포트를 "FactBank(JSON)" 형태로 변환하는 1차 인제스트(전처리) 스크립트입니다.

- PyMuPDF(fitz): 텍스트 블록/좌표 추출
- pdfplumber: 표(table) 추출

주의:
- 본 스크립트는 LLM 없이 규칙 기반으로 "후속 시나리오 생성에 필요한 근거 단위"를 최대한 구조화합니다.
- 정교한 의미 요약/추론이 필요하면, 2단계(LMM/LLM Fact Distiller)를 별도로 두는 것을 권장합니다.

사용 예:
    python factbank_ingest_v2.py --pdf "리포트) 디티앤씨알오_높아지는 실적 턴어라운드 기대감.pdf" --out out.json --analyst-sector-map "analyst_sector_map.xlsx"

옵션:
    --structured-out: 원문 구조(JSON: blocks/tables)도 함께 저장
"""

import argparse
import datetime as _dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import pdfplumber

from openpyxl import load_workbook


# -----------------------------
# Utilities
# -----------------------------
def clean_text(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text).strip()
    return text


def sha256_file(path: str, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            b = f.read(chunk_size)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def _normalize_person_name(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    return re.sub(r"\s+", "", str(name)).strip()

def load_analyst_sector_map(path: Optional[str]) -> Dict[str, str]:
    """
    애널리스트 이름 -> 섹터(coverage_team) 매핑 테이블을 로드합니다.

    지원 포맷:
      - .xlsx: 첫 번째 시트에서 헤더를 자동 탐지 (이름/섹터)
      - .json: {"박유악":"반도체", ...}
      - .csv : name,sector 2컬럼(헤더 포함) 또는 (이름,섹터) 헤더
    """
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"analyst sector map not found: {path}")

    ext = p.suffix.lower()
    if ext == ".json":
        data = json.loads(p.read_text(encoding="utf-8"))
        out = {}
        for k, v in data.items():
            nk = _normalize_person_name(k)
            if nk and v:
                out[nk] = str(v).strip()
        return out

    if ext == ".csv":
        # 아주 단순한 CSV (콤마 구분)만 지원
        out = {}
        with p.open("r", encoding="utf-8") as f:
            rows = [r.strip() for r in f.readlines() if r.strip()]
        if not rows:
            return out
        header = [h.strip().lower() for h in rows[0].split(",")]
        name_i = 0
        sect_i = 1 if len(header) > 1 else 0
        # 헤더 탐지
        for i, h in enumerate(header):
            if h in {"name", "analyst", "애널리스트", "이름", "성명"}:
                name_i = i
            if h in {"sector", "coverage", "team", "섹터", "커버리지", "coverage_team"}:
                sect_i = i
        for r in rows[1:]:
            cols = [c.strip() for c in r.split(",")]
            if len(cols) <= max(name_i, sect_i):
                continue
            nk = _normalize_person_name(cols[name_i])
            sv = cols[sect_i].strip()
            if nk and sv:
                out[nk] = sv
        return out

    # default: xlsx
    wb = load_workbook(filename=str(p), data_only=True)
    ws = wb.worksheets[0]
    # 헤더(첫 3행) 중 가장 그럴듯한 행 찾기
    header_row = None
    header_vals_by_row = {}
    for r in range(1, 4):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        norm = [str(v).strip() if v is not None else "" for v in vals]
        header_vals_by_row[r] = norm
        joined = " ".join(norm).lower()
        if any(k in joined for k in ["이름", "성명", "애널리스트", "name", "analyst"]) and any(
            k in joined for k in ["섹터", "커버리지", "sector", "coverage"]
        ):
            header_row = r
            break
    if header_row is None:
        header_row = 1  # fallback

    headers = [str(ws.cell(row=header_row, column=c).value).strip() if ws.cell(row=header_row, column=c).value is not None else "" 
               for c in range(1, ws.max_column + 1)]
    headers_l = [h.lower() for h in headers]

    name_col = None
    sector_col = None
    for i, h in enumerate(headers_l):
        if h in {"이름", "성명", "애널리스트", "name", "analyst"} or "애널리스트" in h or "이름" in h:
            if name_col is None:
                name_col = i + 1
        if h in {"섹터", "커버리지", "sector", "coverage", "coverage_team", "team"} or "섹터" in h or "커버리지" in h:
            if sector_col is None:
                sector_col = i + 1

    # 최후 fallback: 첫 컬럼=이름, 둘째=섹터
    if name_col is None:
        name_col = 1
    if sector_col is None:
        sector_col = 2

    out = {}
    for r in range(header_row + 1, ws.max_row + 1):
        raw_name = ws.cell(row=r, column=name_col).value
        raw_sector = ws.cell(row=r, column=sector_col).value
        nk = _normalize_person_name(raw_name)
        sv = str(raw_sector).strip() if raw_sector is not None else ""
        if nk and sv:
            out[nk] = sv
    return out


def normalize_date_yyyy_mm_dd(s: str) -> Optional[str]:
    """
    '2025.11.11', '2025. 11. 17' 같은 문자열을 '2025-11-11'로 정규화
    """
    if not s:
        return None
    m = re.search(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})", s)
    if not m:
        return None
    y, mo, d = m.group(1), int(m.group(2)), int(m.group(3))
    return f"{y}-{mo:02d}-{d:02d}"

def extract_numbers(text: str) -> List[Dict[str, Any]]:
    """
    텍스트에서 숫자/단위 후보를 "원문 그대로" 최대한 보존해 추출합니다.
    (후속 검증/표현은 LLM 또는 별도 Validator에서 수행)
    """
    if not text:
        return []

    patterns: List[Tuple[str, str]] = [
        (r"\$ ?[\d,]+(?:\.\d+)?", "usd"),
        (r"[\d,]+(?:\.\d+)?\s*եցին", "percent"),
        (r"[+\- ]\s*[\d,]+(?:\.\d+)?\s*%[A-Za-z]*", "percent_delta"),
        (r"[\d,]+(?:\.\d+)?\s*(?:B|M|K)\b", "short_scale"),
        (r"[\d,]+(?:\.\d+)?\s*(?:억달러|억 달러|백만달러|백만 달러|억 원|억원|조원|조 원|원)\b", "krw_or_unit"),
        (r"\b(?:YoY|QoQ|CY|FY)\b", "period_marker"),
    ]

    seen = set()
    out: List[Dict[str, Any]] = []

    for pat, kind in patterns:
        for m in re.finditer(pat, text):
            raw = m.group(0)
            key = (raw, kind, m.start())
            if key in seen:
                continue
            seen.add(key)
            out.append(
                {
                    "raw": raw,
                    "kind": kind,
                    "start": m.start(),
                    "end": m.end(),
                }
            )

    # 정렬(텍스트 등장 순서)
    out.sort(key=lambda x: x["start"])
    return out

def guess_tags(text: str) -> List[str]:
    """
    단순 키워드 기반 태깅(초기 버전)
    """
    t = text.lower()
    tags = []

    def has_any(keys: List[str]) -> bool:
        return any(k.lower() in t for k in keys)

    if has_any(["목표주가", "투자의견", "컨센서스", "현재주가", "buy", "매수", "보유", "매도"]):
        tags.append("consensus")
    if has_any(["실적", "review", "매출", "영업이익", "순이익", "eps", "마진", "opm", "ebitda"]):
        tags.append("performance")
    if has_any(["가이던스", "guidance", "전망", "전망치", "추정", "예상", "전제", "outlook"]):
        tags.append("outlook")
    if has_any(["리스크", "우려", "불확실", "변동", "경쟁", "규제", "정책"]):
        tags.append("risk")
    if has_any(["전략", "집중", "매각", "투자", "확대", "계약", "파트너", "협업"]):
        tags.append("strategy")
    if has_any(["주:", "자료:", "출처", "기준일", "blumberg", "bloomberg", "gaap", "non-gaap", "ifrs"]):
        tags.append("basis_or_source")
    if has_any(["정의", "란", ":", "용어"]):
        # 너무 광범위하지만, glossary 후보를 추가로 거를 때 사용
        tags.append("definition_candidate")

    # 중복 제거
    tags = list(dict.fromkeys(tags))
    return tags

def is_candidate_fact(text: str, page: int) -> bool:
    """
    FactBank에 넣을 '근거 문장' 후보인지 결정.
    - 페이지 1은 상대적으로 넓게 포함
    - 그 외 페이지는 키워드/숫자 포함 중심으로 선별
    """
    if not text or len(text) < 15:
        return False

    # 페이지 번호/잡음 제거(너무 짧은 숫자만)
    if re.fullmatch(r"[\d\.\-]+", text):
        return False

    # 1페이지는 핵심이 많으므로 완화
    if page == 1:
        return True if len(text) >= 20 else False

    # 다른 페이지는 숫자나 핵심 키워드가 있을 때만
    keywords = ["실적", "가이던스", "컨센서스", "전망", "목표주가", "매출", "EPS", "YoY", "QoQ", "리스크", "전략", "투자"]
    if any(k.lower() in text.lower() for k in keywords):
        return True
    if re.search(r"\$ ?[\d,]+(?:\.\d+)?|[\d,]+(?:\.\d+)?\s*եցին", text):
        return True

    return False


# -----------------------------
# PDF Extraction (Structured)
# -----------------------------
def extract_structured(pdf_path: str) -> Dict[str, Any]:
    doc_data: Dict[str, Any] = {
        "doc_meta": {
            "title": pdf_path,
            "page_count": 0,
            "status": "ingested",
        },
        "blocks": [],
        "tables": [],
    }

    # A) Tables via pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        doc_data["doc_meta"]["page_count"] = len(pdf.pages)
        for p_idx, page in enumerate(pdf.pages):
            page_num = p_idx + 1
            try:
                tables = page.extract_tables()
            except Exception:
                tables = []

            for t_idx, table in enumerate(tables or []):
                cleaned_matrix = [
                    [clean_text(str(cell)) if cell else "" for cell in row]
                    for row in table
                ]
                table_id = f"t_{page_num}_{t_idx+1}"
                doc_data["tables"].append(
                    {
                        "id": table_id,
                        "page": page_num,
                        "matrix": cleaned_matrix,
                        "raw_text": clean_text(str(cleaned_matrix)),
                    }
                )

    # B) Text blocks via PyMuPDF
    doc = fitz.open(pdf_path)
    for p_idx, page in enumerate(doc):
        page_num = p_idx + 1
        blocks = page.get_text("dict").get("blocks", [])
        for b_idx, block in enumerate(blocks):
            b_type = block.get("type")

            # 0: text block
            if b_type == 0:
                block_text = ""
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        block_text += span.get("text", "") + " "

                block_text = clean_text(block_text)
                if len(block_text) < 5:
                    continue

                doc_data["blocks"].append(
                    {
                        "id": f"p{page_num}_b{b_idx}",
                        "page": page_num,
                        "type": "text",
                        "text": block_text,
                        "bbox": block.get("bbox"),
                    }
                )
                continue

            # 1: image block (차트/스크린샷 등)
            if b_type == 1:
                doc_data["blocks"].append(
                    {
                        "id": f"p{page_num}_img{b_idx}",
                        "page": page_num,
                        "type": "image",
                        "bbox": block.get("bbox"),
                        # 원본 바이트는 저장하지 않음(용량 급증). 필요 시 후속 단계에서 bbox로 crop.
                    }
                )
                continue

            # 기타 타입은 무시
            continue


    doc.close()
    return doc_data


# -----------------------------
# Report Meta Extraction
# -----------------------------
def extract_report_meta(structured: Dict[str, Any], analyst_sector_map: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
    """
    1페이지 텍스트를 기반으로 report_meta를 최대한 채웁니다.
    - PDF 추출 텍스트는 블록 단위라 줄바꿈이 완벽하지 않을 수 있으므로,
      가능한 범위에서 "라인 기반"으로 우선 탐색하고, 실패 시 전체 텍스트에서 fallback합니다.
    """
    blocks = [b for b in structured.get("blocks", []) if b.get("page") == 1]

    # 1페이지 전체 텍스트(대략적인 라인 재구성)
    page1_text = "\n".join([b.get("text", "") for b in blocks])
    lines = [clean_text(x) for x in page1_text.splitlines() if clean_text(x)]

    # 날짜
    report_date = normalize_date_yyyy_mm_dd(page1_text)

    # 이메일
    email = None
    m = re.search(r"\b[\w\.-]+@[\w\.-]+\b", page1_text)
    if m:
        email = m.group(0)

    # 애널리스트 이름
    author_name = None
    # "반도체 Analyst 박유악", "US Strategy 김승혁" 등
    m = re.search(r"(?:Analyst|애널리스트)\s*([가-힣]{2,4})", page1_text)
    if m:
        author_name = m.group(1)
    else:
        # 이메일 앞에 이름이 붙는 경우를 보완
        if email:
            m = re.search(r"([가-힣]{2,4})\s+" + re.escape(email), page1_text)
            if m:
                author_name = m.group(1)

    # 회사명/티커(라인 우선)
    company = None
    ticker = None
    company_line_idx: Optional[int] = None
    comp_pat = re.compile(r"^(.+?)\s*\(([A-Z0-9]{1,6}\.US|\d{6})\)$")

    for i, line in enumerate(lines):
        if "@" in line:
            continue
        m = comp_pat.search(line)
        if m:
            company = clean_text(m.group(1))
            ticker = clean_text(m.group(2))
            company_line_idx = i
            break

    # fallback: 전체 텍스트에서 찾되 이메일/센터 문구가 섞이지 않도록 정리
    if not (company and ticker):
        scrubbed = re.sub(r"\b[\w\.-]+@[\w\.-]+\b", " ", page1_text)  # 이메일 제거
        scrubbed = re.sub(r"키움증권.*?\|", " ", scrubbed)  # 헤더 일부 제거(과매칭 방지)
        m = re.search(r"([A-Za-z가-힣0-9&·\.\- ]{2,40})\s*\(([A-Z0-9]{1,6}\.US|\d{6})\)", scrubbed)
        if m:
            company = clean_text(m.group(1))
            ticker = clean_text(m.group(2))

    # 제목: 회사 라인 다음 줄을 우선
    title = None
    if company_line_idx is not None:
        # 다음 라인을 title로 시도
        if company_line_idx + 1 < len(lines):
            title = lines[company_line_idx + 1]
        # title이 지나치게 일반적인 경우(예: "Stock Data")는 다음 줄로 보정
        if title and title.lower() in {"stock data", "earnings & valuation", "performance & price trend"}:
            if company_line_idx + 2 < len(lines):
                title = lines[company_line_idx + 2]

    # 커버리지(coverage_team): '조직/데스크'가 아니라 애널리스트의 담당 섹터를 우선으로 채웁니다.
    analyst_sector_map = analyst_sector_map or {}
    cov_from_map = False
    coverage_team = None

    # 0) (권장) 매핑 테이블: 애널리스트 이름 -> 섹터
    if author_name:
        mapped = analyst_sector_map.get(_normalize_person_name(author_name))
        if mapped:
            coverage_team = str(mapped).strip()
            cov_from_map = True

    # 1) type2 헤더: "키움증권리서치센터 | 미국기업분석" 같은 포맷(fallback)
    if not coverage_team:
        m = re.search(r"키움증권\s*리서치센터\s*\|\s*([^\n]+)", page1_text)
        if not m:
            m = re.search(r"키움증권리서치센터\s*\|\s*([^\n]+)", page1_text)  # 붙어있는 케이스
        if m:
            coverage_team = clean_text(m.group(1))

    # 2) type1: "의료기기 Analyst 신민수"처럼 섹터명이 Analyst 앞에 붙는 케이스(fallback)
    if not coverage_team:
        m = re.search(r"([가-힣A-Za-z&/· ]{2,30})\s*(?:Analyst|애널리스트)\s*[가-힣]{2,4}", page1_text)
        if m:
            coverage_team = clean_text(m.group(1))

    return {
        "title": title,
        "company": company,
        "ticker": ticker,
        "report_date": report_date,
        "author_name": author_name,
        "coverage_team": coverage_team,
        "language": "ko",
        "meta_confidence": {
            "title": 0.6 if title else 0.0,
            "company_ticker": 0.8 if company and ticker else 0.0,
            "report_date": 0.8 if report_date else 0.0,
            "author_name": 0.7 if author_name else 0.0,
            "coverage_team": 0.95 if cov_from_map else (0.7 if coverage_team else 0.0),
        },
    }

def extract_highlights(structured: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    1페이지에서 '' 또는 '•' 등의 하이라이트를 추출
    """
    blocks = [b for b in structured.get("blocks", []) if b.get("page") == 1]
    highlights: List[Dict[str, Any]] = []

    # 1페이지에서 bullet로 시작하는 문장을 잡는다
    bullet_pat = re.compile(r"(?:^| )([•\-]\s*[^•\-]+)")

    for b in blocks:
        t = b.get("text", "")
        # PDF 추출 텍스트는 줄바꿈이 사라져 bullet 구분이 어려울 수 있어, 기호 기반으로 찾음
        matches = re.findall(r"[•]\s*[^•]+", t)
        for m in matches:
            item = clean_text(m.replace("", "").replace("•", ""))
            if item and len(item) >= 8:
                highlights.append(
                    {
                        "text": item,
                        "evidence": {"page": 1, "source_type": "text", "source_id": b["id"], "excerpt": item[:180]},
                    }
                )

    # 중복 제거
    uniq = []
    seen = set()
    for h in highlights:
        key = h["text"]
        if key in seen:
            continue
        seen.add(key)
        uniq.append(h)

    return uniq

def extract_footnotes(structured: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    '주:' '자료:' 같은 문구를 footnotes로 모음
    """
    out: List[Dict[str, Any]] = []
    for b in structured.get("blocks", []):
        t = b.get("text", "")
        if "주:" in t or "자료:" in t or "출처" in t:
            # 너무 길면 일부만
            excerpt = t[:260]
            out.append(
                {
                    "text": t,
                    "evidence": {"page": b["page"], "source_type": "text", "source_id": b["id"], "excerpt": excerpt},
                }
            )
    return out


# -----------------------------
# FactBank Builder
# -----------------------------
def build_factbank(pdf_path: str, structured: Dict[str, Any], structured_out: Optional[str] = None,
                   mode: str = "compact", max_facts: int = 45, max_tables: int = 6, max_highlights: int = 8, analyst_sector_map: Optional[Dict[str, str]] = None, include_pdf_sha256: bool = False) -> Dict[str, Any]:
    report_meta = extract_report_meta(structured, analyst_sector_map=analyst_sector_map)
    highlights = extract_highlights(structured)
    footnotes = extract_footnotes(structured)

    facts: List[Dict[str, Any]] = []
    fid = 1

    # highlight를 fact로도 등록(검색 편의)
    for h in highlights:
        facts.append(
            {
                "fact_id": f"F{fid:04d}",
                "claim": h["text"],
                "tags": ["highlight"],
                "numbers": extract_numbers(h["text"]),
                "evidence": h["evidence"],
            }
        )
        fid += 1

    # 텍스트 블록 기반 fact 후보 생성
    for b in structured.get("blocks", []):
        page = int(b.get("page", 0))
        text = b.get("text", "")
        if not is_candidate_fact(text, page):
            continue

        # 너무 긴 문단은 잘라서 evidence excerpt로만 두고 claim은 축약본을 만든다
        claim = text
        if len(claim) > 450:
            claim = claim[:450] + "…"

        facts.append(
            {
                "fact_id": f"F{fid:04d}",
                "claim": claim,
                "tags": guess_tags(text),
                "numbers": extract_numbers(text),
                "evidence": {
                    "page": page,
                    "source_type": "text",
                    "source_id": b.get("id"),
                    "excerpt": text[:260],
                },
            }
        )
        fid += 1

    # 표를 table 섹션으로 저장(시나리오에서 수치 자막 생성 시 활용)
    tables_out: List[Dict[str, Any]] = []
    for t in structured.get("tables", []):
        matrix = t.get("matrix", [])
        raw_text = t.get("raw_text", "")
        tables_out.append(
            {
                "table_id": t.get("id"),
                "page": t.get("page"),
                "row_count": len(matrix),
                "col_count": max((len(r) for r in matrix), default=0),
                "matrix": matrix,
                "numbers": extract_numbers(raw_text),
                "tags": ["table"],
                "evidence": {
                    "page": t.get("page"),
                    "source_type": "table",
                    "source_id": t.get("id"),
                    "excerpt": raw_text[:260],
                },
            }
        )



    # 이미지/차트 등 figure 메타(후속 단계에서 bbox로 crop 가능)
    figures_out: List[Dict[str, Any]] = []
    for b in structured.get("blocks", []):
        if b.get("type") == "image":
            figures_out.append(
                {
                    "figure_id": b.get("id"),
                    "page": int(b.get("page", 0) or 0),
                    "bbox": b.get("bbox"),
                    "tags": ["figure"],
                    "evidence": {
                        "page": int(b.get("page", 0) or 0),
                        "source_type": "image",
                        "source_id": b.get("id"),
                        "excerpt": "image_block",
                    },
                }
            )


    # ---- Size control (token economy) ----
    # highlights는 상단만 유지
    if max_highlights is not None and max_highlights > 0:
        highlights = highlights[:max_highlights]

    def _page_weight(p: int) -> int:
        if p <= 0:
            return 0
        if p == 1:
            return 6
        if p == 2:
            return 4
        if p == 3:
            return 3
        return 1

    def _score_fact(f: Dict[str, Any]) -> float:
        ev = f.get("evidence", {}) or {}
        p = int(ev.get("page", 0) or 0)
        score = _page_weight(p)

        tags = set(f.get("tags", []) or [])
        tag_w = {
            "highlight": 6,
            "consensus": 5,
            "performance": 5,
            "outlook": 4,
            "risk": 4,
            "strategy": 3,
            "basis_or_source": 2,
        }
        score += sum(tag_w.get(t, 0) for t in tags)

        nums = f.get("numbers", []) or []
        score += min(len(nums), 6) * 0.6

        claim = f.get("claim", "") or ""
        # 너무 긴 문단은 가점 없음(후속에서 요약 부담)
        if 30 <= len(claim) <= 220:
            score += 1.0
        return score

    def _score_table(t: Dict[str, Any]) -> float:
        p = int(t.get("page", 0) or 0)
        score = _page_weight(p)
        # 작은 테이블은 유용(장면 자막으로 쓰기 쉬움)
        rc = int(t.get("row_count", 0) or 0)
        cc = int(t.get("col_count", 0) or 0)
        size = rc * cc
        if 0 < size <= 300:
            score += 2.0
        elif 300 < size <= 900:
            score += 1.0
        else:
            score += 0.0
        score += min(len(t.get("numbers", []) or []), 10) * 0.2
        return score

    if mode == "compact":
        # highlight fact는 항상 유지
        highlight_facts = [f for f in facts if "highlight" in (f.get("tags") or [])]
        other_facts = [f for f in facts if "highlight" not in (f.get("tags") or [])]
        other_facts.sort(key=_score_fact, reverse=True)

        if max_facts is not None and max_facts > 0:
            other_facts = other_facts[:max_facts]

        facts = highlight_facts + other_facts

        # tables도 상위만 유지
        tables_out.sort(key=_score_table, reverse=True)
        if max_tables is not None and max_tables > 0:
            tables_out = tables_out[:max_tables]

        # figures도 너무 많으면 상단만 유지
        if len(figures_out) > 10:
            figures_out = figures_out[:10]

        # footnotes도 길이/개수 제어
        if len(footnotes) > 20:
            footnotes = footnotes[:20]

    now = _dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    fb_meta = {
        "schema_version": "0.1",
        "generated_at": now,
        "generator": "factbank_ingest.py",
        "source_pdf": pdf_path,
        "page_count": structured.get("doc_meta", {}).get("page_count"),
        "status": "ok",
    }
    if include_pdf_sha256:
        fb_meta["source_pdf_sha256"] = sha256_file(pdf_path)

    fb = {
        "factbank_meta": fb_meta,
        "report_meta": report_meta,
        "highlights": highlights,
        "facts": facts,
        "tables": tables_out,
        "figures": figures_out,
        "footnotes": footnotes,
        "raw_ref": {
            "structured_json": structured_out,
        },
    }

    return fb

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", required=True, help="입력 PDF 경로")
    ap.add_argument("--out", required=True, help="FactBank JSON 출력 경로")
    ap.add_argument("--structured-out", default=None, help="(선택) blocks/tables 구조 JSON 출력 경로")
    ap.add_argument("--mode", choices=["compact", "full"], default="compact", help="출력 크기 제어 모드")
    ap.add_argument("--max-facts", type=int, default=45, help="(compact 모드) 최대 facts 개수")
    ap.add_argument("--max-tables", type=int, default=6, help="(compact 모드) 최대 tables 개수")
    ap.add_argument("--max-highlights", type=int, default=8, help="최대 highlights 개수")
    ap.add_argument("--analyst-sector-map", default=None, help="(선택) 애널리스트 이름→섹터 매핑 파일(.xlsx/.csv/.json)")
    ap.add_argument("--include-pdf-sha256", action="store_true", help="(선택) source_pdf_sha256를 factbank_meta에 포함")
    args = ap.parse_args()

    analyst_sector_map = load_analyst_sector_map(args.analyst_sector_map) if args.analyst_sector_map else {}

    structured = extract_structured(args.pdf)

    structured_out = None
    if args.structured_out:
        structured_out = args.structured_out
        with open(structured_out, "w", encoding="utf-8") as f:
            json.dump(structured, f, ensure_ascii=False, indent=2)

    factbank = build_factbank(
        args.pdf,
        structured,
        structured_out=structured_out,
        mode=args.mode,
        max_facts=args.max_facts,
        max_tables=args.max_tables,
        max_highlights=args.max_highlights,
        analyst_sector_map=analyst_sector_map,
        include_pdf_sha256=args.include_pdf_sha256,
    )

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(factbank, f, ensure_ascii=False, indent=2)

    print("✅ FactBank 생성 완료")
    print(f"- PDF: {args.pdf}")
    print(f"- FactBank: {args.out}")
    if structured_out:
        print(f"- Structured: {structured_out}")
    print(f"- mode: {args.mode}")
    print(f"- facts: {len(factbank['facts'])}, tables: {len(factbank['tables'])}")


if __name__ == "__main__":
    main()


# --- API Wrapper for App Integration ---
def extract_facts_from_pdf(pdf_path: str) -> Dict[str, Any]:
    """
    App integration wrapper.
    Takes a PDF file path, runs the ingestion pipeline, and returns the FactBank dict.
    """
    # 1. Extract structure (blocks/tables)
    structured = extract_structured(pdf_path)
    
    # 2. Build FactBank
    # Defaulting to 'compact' mode and default limits as per CLI defaults
    factbank = build_factbank(
        pdf_path=pdf_path,
        structured=structured,
        mode="compact",
        max_facts=45,
        max_tables=6,
        max_highlights=8,
        analyst_sector_map=None, # Or load default map if needed
        include_pdf_sha256=False
    )
    
    return factbank
